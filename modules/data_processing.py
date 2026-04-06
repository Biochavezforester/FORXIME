"""
Módulo de procesamiento de datos para FORXIME/2
"""
import pandas as pd
import numpy as np
from datetime import datetime
import streamlit as st
import io
from openpyxl import load_workbook
from utils.validators import (validate_utm_coordinates, validate_date, 
                              validate_time, validate_excel_format)
from utils.geospatial import group_cameras_by_distance, add_latlon_columns
from utils.helpers import calculate_independent_events, clean_species_name


def create_excel_template():
    """
    Crea un DataFrame de ejemplo para la plantilla Excel
    
    Returns:
        DataFrame: Plantilla con datos de ejemplo
    """
    template_data = {
        'Sitio': ['Sitio_A', 'Sitio_A', 'Sitio_B', 'Sitio_B'],
        'Camara': ['CAM001', 'CAM001', 'CAM002', 'CAM002'],
        'Coordenada_X_UTM': [500000, 500000, 501000, 501000],
        'Coordenada_Y_UTM': [2500000, 2500000, 2501000, 2501000],
        'Zona_UTM': ['12N', '12N', '12N', '12N'],
        'Especie_Categoria': ['Panthera onca', 'Tapirus bairdii', 'Panthera onca', 'Pecari tajacu'],
        'Fecha': ['01/01/2026', '02/01/2026', '01/01/2026', '03/01/2026'],
        'Hora': ['08:30:00', '14:45:00', '22:15:00', '06:20:00'],
        'Eventos_Independientes': [1, 1, 1, 2],
        'Es_Cria': ['No', 'No', 'No', 'Sí'],
        'Lactante': ['No', 'No', 'No', 'No'],
        'Periodo_Ensenanza': ['No', 'No', 'No', 'Sí'],
        'Rascando_Arboles': ['No', 'No', 'Sí', 'No'],
        'Usando_Letrina': ['No', 'No', 'No', 'No'],
        'Salud_Fisica': ['Buena', 'Buena', 'Buena', 'Buena'],
        'Observaciones': ['', '', 'Marcaje territorial', 'Grupo familiar']
    }
    
    return pd.DataFrame(template_data)


def process_manual_data(manual_data):
    """
    Procesa datos ingresados manualmente
    
    Args:
        manual_data: Lista de diccionarios con datos manuales
    
    Returns:
        DataFrame: Datos procesados
    """
    if not manual_data:
        return None
    
    df = pd.DataFrame(manual_data)
    
    # Limpiar nombres de especies
    if 'Especie_Categoria' in df.columns:
        df['Especie_Categoria'] = df['Especie_Categoria'].apply(clean_species_name)
    
    # Agregar columnas de lat/lon
    df = add_latlon_columns(df)
    
    return df


def process_excel_data(uploaded_file):
    """
    Procesa archivo Excel cargado con detección automática de entorno para máxima velocidad.
    """
    import sys
    
    # DETECCIÓN DE ENTORNO: Usar motor nativo de pandas si es Windows (Local/Portable)
    # En Wasm (stlite/browser), sys.platform suele ser 'emscripten' o 'pyodide'
    is_browser = sys.platform in ['emscripten', 'pyodide']
    
    if not is_browser:
        # LECTOR NUCLEAR (Nativo): 100x más rápido que pd.read_excel en archivos grandes
        try:
            if hasattr(uploaded_file, 'read'):
                file_content = io.BytesIO(uploaded_file.read())
            else:
                file_content = uploaded_file

            # CARGA DE SOLO LECTURA (Sin estilos ni basura de Excel)
            wb = load_workbook(file_content, read_only=True, data_only=True)
            sheet = wb.active
            
            data = []
            # Límite de 50 columnas para evitar escanear el "infinito" de Excel
            max_col = min(sheet.max_column, 50) if sheet.max_column else 50
            
            for row in sheet.iter_rows(max_col=max_col, values_only=True):
                # Si llegamos a una fila totalmente vacía, terminamos (SÚPER RÁPIDO)
                if not any(row):
                    if len(data) > 0: break
                    continue
                data.append(row)
                
            if not data:
                return False, "El archivo Excel está vacío."
            
            # Convertimos a DataFrame (Ahora sí es instantáneo)
            df = pd.DataFrame(data[1:], columns=data[0])
            wb.close()
            
        except Exception as e:
            return False, f"Error en lectura nuclear: {str(e)}"
    else:
        # LECTOR WEB (Wasm): Solo para el navegador
        try:
            if hasattr(uploaded_file, 'read'):
                file_content = io.BytesIO(uploaded_file.read())
            else:
                file_content = uploaded_file

            wb = load_workbook(file_content, read_only=True, data_only=True)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(values_only=True):
                if not any(row): break
                data.append(row)
                
            if not data: return False, "El archivo Excel está vacío."
            
            df = pd.DataFrame(data[1:], columns=data[0])
            wb.close()
        except Exception as e:
            return False, f"Error en lectura Web: {str(e)}"

    # PROCESAMIENTO COMÚN (Minimalista y Rápido)
    try:
        # Solo limpieza básica de nulos
        df = df.fillna(np.nan)
        
        # Validar formato estructural (Solo columnas)
        is_valid, result = validate_excel_format(df)
        if not is_valid: return False, result
        
        # NORMALIZACIÓN FÉRREA DE TIPOS (Evita errores de comparación str vs datetime)
        if 'Fecha' in df.columns:
            df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
        
        if 'Hora' in df.columns:
            # Asegurar que Hora sea consistente (time object)
            df['Hora'] = pd.to_datetime(df['Hora'].astype(str), format='mixed', errors='coerce').dt.time
            
        if 'Especie_Categoria' in df.columns:
            df['Especie_Categoria'] = df['Especie_Categoria'].astype(str).str.strip().str.title()
        
        # Eliminar filas con fechas inválidas que podrían romper los cálculos de .min()/.max()
        if 'Fecha' in df.columns:
            df = df.dropna(subset=['Fecha'])
        
        return True, df
    
    except Exception as e:
        return False, f"Error procesando archivo Excel: {str(e)}"


def group_sites(df, max_distance=10):
    """
    Agrupa sitios basado en distancia entre cámaras
    
    Args:
        df: DataFrame con datos de cámaras
        max_distance: Distancia máxima en metros para agrupar
    
    Returns:
        DataFrame: DataFrame con sitios agrupados
    """
    # Obtener cámaras únicas
    cameras_unique = df[['Camara', 'Coordenada_X_UTM', 'Coordenada_Y_UTM', 'Zona_UTM']].drop_duplicates()
    
    # Agrupar cámaras
    cameras_grouped = group_cameras_by_distance(cameras_unique, max_distance)
    
    # Merge con datos originales
    df = df.merge(cameras_grouped[['Camara', 'Sitio_Agrupado']], on='Camara', how='left')
    
    return df


def prepare_detection_history(df):
    """
    Prepara matriz de historia de detección para análisis de ocupación
    
    Args:
        df: DataFrame con datos procesados
    
    Returns:
        dict: Diccionario con matrices de detección por especie
    """
    detection_histories = {}
    
    # Convertir fecha a datetime
    df['Fecha'] = pd.to_datetime(df['Fecha'])
    
    # Definir ocasiones de muestreo (por ejemplo, semanas)
    df['Semana'] = df['Fecha'].dt.isocalendar().week
    
    for species in df['Especie_Categoria'].unique():
        species_data = df[df['Especie_Categoria'] == species]
        
        # Crear matriz de detección (sitios x ocasiones)
        detection_matrix = species_data.pivot_table(
            index='Sitio_Agrupado',
            columns='Semana',
            values='Eventos_Independientes',
            aggfunc='sum',
            fill_value=0
        )
        
        # Convertir a binario (1 si hubo detección, 0 si no)
        detection_binary = (detection_matrix > 0).astype(int)
        
        detection_histories[species] = {
            'counts': detection_matrix,
            'binary': detection_binary
        }
    
    return detection_histories


def calculate_basic_metrics(df):
    """
    Calcula métricas básicas del dataset
    
    Args:
        df: DataFrame con datos procesados
    
    Returns:
        dict: Diccionario con métricas básicas
    """
    metrics = {
        'total_records': len(df),
        'total_cameras': df['Camara'].nunique(),
        'total_sites': df['Sitio_Agrupado'].nunique() if 'Sitio_Agrupado' in df.columns else df['Sitio'].nunique(),
        'total_species': df['Especie_Categoria'].nunique(),
        'date_range': {
            'start': df['Fecha'].min() if not df['Fecha'].empty else None,
            'end': df['Fecha'].max() if not df['Fecha'].empty else None,
            'days': (pd.to_datetime(df['Fecha'].max()) - pd.to_datetime(df['Fecha'].min())).days if not df['Fecha'].empty else 0
        },
        'total_independent_events': df['Eventos_Independientes'].sum() if 'Eventos_Independientes' in df.columns else len(df)
    }
    
    return metrics


def identify_non_wildlife(df):
    """
    Identifica registros que no son fauna silvestre
    
    Args:
        df: DataFrame con datos
    
    Returns:
        DataFrame: Registros no-fauna silvestre
    """
    non_wildlife_keywords = [
        'humano', 'human', 'persona', 'person', 'gente', 'people',
        'perro', 'dog', 'can', 'domestico', 'domestic',
        'gato', 'cat', 'feline domestic',
        'vaca', 'cow', 'ganado', 'cattle', 'bovino',
        'caballo', 'horse', 'equino',
        'vehiculo', 'vehicle', 'carro', 'car', 'moto', 'motorcycle',
        'bicicleta', 'bicycle', 'bike'
    ]
    
    # Crear máscara para identificar no-fauna
    # Usar \b para coincidir con la palabra completa y evitar falsos positivos (ej: 'can' en 'canario')
    keywords_regex = r'\b(' + '|'.join(non_wildlife_keywords).replace('|', r'|') + r')\b'
    mask = df['Especie_Categoria'].str.lower().str.contains(keywords_regex, regex=True, na=False)
    
    non_wildlife_df = df[mask].copy()
    non_wildlife_df['Categoria_Antropogenica'] = non_wildlife_df['Especie_Categoria'].apply(
        lambda x: categorize_anthropogenic(x)
    )
    
    return non_wildlife_df


def categorize_anthropogenic(species_name):
    """
    Categoriza registros antropogénicos
    
    Args:
        species_name: Nombre de la especie/categoría
    
    Returns:
        str: Categoría antropogénica
    """
    import re
    species_lower = str(species_name).lower()
    
    # Usamos \b para asegurar que se busque la palabra completa y no un fragmento (ej: evitar car en carpintero)
    if re.search(r'\b(humano|human|persona|person|gente|people)\b', species_lower):
        return 'Humano'
    elif re.search(r'\b(perro|dog)\b', species_lower):
        return 'Perro Doméstico'
    elif re.search(r'\b(gato|cat)\b', species_lower):
        return 'Gato Doméstico'
    elif re.search(r'\b(vaca|cow|ganado|cattle|bovino)\b', species_lower):
        return 'Ganado'
    elif re.search(r'\b(vehiculo|vehículo|vehicle|carro|car|moto)\b', species_lower):
        return 'Vehículo'
    else:
        return 'Otro Antropogénico'


def filter_wildlife_only(df):
    """
    Filtra solo registros de fauna silvestre
    
    Args:
        df: DataFrame con todos los datos
    
    Returns:
        DataFrame: Solo fauna silvestre
    """
    non_wildlife = identify_non_wildlife(df)
    wildlife_df = df[~df.index.isin(non_wildlife.index)].copy()
    
    return wildlife_df


def add_behavior_data(df, behavior_records):
    """
    Agrega datos de comportamiento al DataFrame
    
    Args:
        df: DataFrame principal
        behavior_records: Lista de diccionarios con observaciones de comportamiento
    
    Returns:
        DataFrame: DataFrame con datos de comportamiento agregados
    """
    if not behavior_records:
        return df
    
    behavior_df = pd.DataFrame(behavior_records)
    
    # Merge con DataFrame principal
    # Asumiendo que behavior_records tiene identificadores únicos para cada registro
    df = df.merge(behavior_df, how='left', left_index=True, right_on='record_id')
    
    return df


def summarize_by_camera(df):
    """
    Resume datos por cámara
    
    Args:
        df: DataFrame con datos
    
    Returns:
        DataFrame: Resumen por cámara
    """
    summary = df.groupby('Camara').agg({
        'Especie_Categoria': 'nunique',
        'Eventos_Independientes': 'sum',
        'Fecha': lambda x: (x.max() - x.min()).days + 1
    }).reset_index()
    
    summary.columns = ['Camara', 'Riqueza_Especies', 'Total_Eventos', 'Dias_Activa']
    
    # Calcular tasa de captura
    summary['Tasa_Captura_100dias'] = (summary['Total_Eventos'] / summary['Dias_Activa']) * 100
    
    return summary


def summarize_by_species(df):
    """
    Resume datos por especie
    
    Args:
        df: DataFrame con datos
    
    Returns:
        DataFrame: Resumen por especie
    """
    summary = df.groupby('Especie_Categoria').agg({
        'Camara': 'nunique',
        'Eventos_Independientes': 'sum',
        'Sitio_Agrupado': 'nunique' if 'Sitio_Agrupado' in df.columns else 'Sitio'
    }).reset_index()
    
    summary.columns = ['Especie', 'N_Camaras', 'Total_Eventos', 'N_Sitios']
    
    # Calcular ocupación naive
    total_sites = df['Sitio_Agrupado'].nunique() if 'Sitio_Agrupado' in df.columns else df['Sitio'].nunique()
    summary['Ocupacion_Naive'] = summary['N_Sitios'] / total_sites
    
    # Ordenar por total de eventos
    summary = summary.sort_values('Total_Eventos', ascending=False)
    
    return summary
